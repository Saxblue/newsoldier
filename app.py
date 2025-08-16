import streamlit as st
import pandas as pd
import json
import os
from datetime import datetime, date, timedelta
import plotly.express as px
import plotly.graph_objects as go
import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO
import traceback
import calendar

# Sayfa konfigürasyonu
st.set_page_config(
    page_title="CashBack Düzeltmesi Analizi",
    page_icon="💰",
    layout="wide",
    initial_sidebar_state="expanded"
)

class ExcelProcessor:
    def __init__(self):
        self.required_columns = ['Kullanıcı ID', 'Kullanıcı Adı', 'Bonus Türü']
        self.cashback_value = "CashBack Düzeltmesi"
    
    def process_cashback_data(self, df):
        """Excel dosyasından CashBack Düzeltmesi verilerini işler"""
        try:
            # Sütun isimlerini normalize et
            df.columns = df.columns.str.strip()
            
            st.info(f"📋 Bulunan sütunlar: {list(df.columns)}")
            
            # Sütun eşleştirmesi
            column_mapping = {}
            
            # ID sütununu bul - öncelik sırasına göre
            possible_id_columns = ['müşteri kimliği', 'müşteri_kimliği', 'kullanıcı id', 'kullanıcı_id', 'customer_id']
            for possible in possible_id_columns:
                for col in df.columns:
                    if possible in col.lower():
                        column_mapping['ID'] = col
                        break
                if 'ID' in column_mapping:
                    break
            
            # Eğer bulunamadıysa B sütununu kontrol et
            if 'ID' not in column_mapping:
                if len(df.columns) > 1:
                    column_mapping['ID'] = df.columns[1]
            
            # İsim sütununu bul
            possible_name_columns = ['kullanıcı adı', 'kullanıcı_adı', 'müşteri adı', 'müşteri_adı', 'ad', 'isim']
            for col in df.columns:
                if any(possible in col.lower() for possible in possible_name_columns):
                    column_mapping['Ad'] = col
                    break
            
            # Miktar sütununu bul
            possible_amount_columns = ['para birimi miktar', 'miktar', 'tutar', 'amount', 'toplam']
            for col in df.columns:
                if any(possible in col.lower() for possible in possible_amount_columns):
                    column_mapping['Miktar'] = col
                    break
            
            # Gerekli sütunların kontrolü
            if 'ID' not in column_mapping:
                st.error("❌ Müşteri ID sütunu bulunamadı!")
                return pd.DataFrame()
            
            if 'Ad' not in column_mapping:
                st.error("❌ Müşteri Adı sütunu bulunamadı!")
                return pd.DataFrame()
            
            # Sütun eşleştirmesini göster
            st.success("✅ Sütun eşleştirmesi tamamlandı:")
            for key, value in column_mapping.items():
                st.write(f"   - {key} → {value}")
            
            # Tüm satırlar CashBack Düzeltmesi olduğu için direkt işleme devam et
            cashback_df = df.copy()
            
            # Boş satırları temizle
            cashback_df = cashback_df.dropna(subset=[column_mapping['ID'], column_mapping['Ad']])
            
            if cashback_df.empty:
                st.warning("⚠️ İşlenebilir veri bulunamadı!")
                return pd.DataFrame()
            
            st.success(f"✅ {len(cashback_df)} adet CashBack Düzeltmesi kaydı bulundu!")
            
            # Müşteri bazında gruplama yap
            id_col = column_mapping['ID']
            name_col = column_mapping['Ad']
            
            if 'Miktar' in column_mapping:
                amount_col = column_mapping['Miktar']
                # Miktar sütununu sayısal veriye çevir
                cashback_df[amount_col] = pd.to_numeric(cashback_df[amount_col], errors='coerce').fillna(0)
                
                # Her müşteri için işlem sayısı ve toplam miktar
                count_data = cashback_df.groupby([id_col, name_col]).size().reset_index(name='Adet')
                sum_data = cashback_df.groupby([id_col, name_col])[amount_col].sum().reset_index()
                
                # İki DataFrame'i birleştir
                grouped = count_data.merge(sum_data, on=[id_col, name_col])
                
                # Sütun isimlerini düzenle
                grouped = grouped.rename(columns={
                    id_col: 'Müşteri_Kimliği', 
                    name_col: 'Müşteri_Adı',
                    amount_col: 'Toplam_Miktar'
                })
            else:
                # Miktar sütunu yoksa sadece işlem sayısı
                grouped = cashback_df.groupby([id_col, name_col]).size().reset_index(name='Adet')
                grouped['Toplam_Miktar'] = 0
                
                # Sütun isimlerini düzenle
                grouped = grouped.rename(columns={id_col: 'Müşteri_Kimliği', name_col: 'Müşteri_Adı'})
            
            # Veri tiplerini düzelt
            grouped['Müşteri_Kimliği'] = pd.to_numeric(grouped['Müşteri_Kimliği'], errors='coerce')
            grouped['Adet'] = pd.to_numeric(grouped['Adet'], errors='coerce').fillna(0).astype(int)
            grouped['Toplam_Miktar'] = pd.to_numeric(grouped['Toplam_Miktar'], errors='coerce').fillna(0)
            
            # NaN değerleri temizle
            grouped = grouped.dropna(subset=['Müşteri_Kimliği'])
            
            # En yüksek miktardan en düşüğe sırala
            if grouped['Toplam_Miktar'].sum() > 0:
                grouped = grouped.sort_values('Toplam_Miktar', ascending=False).reset_index(drop=True)
            else:
                grouped = grouped.sort_values('Adet', ascending=False).reset_index(drop=True)
            
            st.success(f"✅ {len(grouped)} farklı müşterinin CashBack analizi tamamlandı!")
            
            return grouped
            
        except Exception as e:
            st.error(f"❌ Veri işleme hatası: {str(e)}")
            st.error(f"📋 Hata detayı: {traceback.format_exc()}")
            return pd.DataFrame()
    
    def create_formatted_excel(self, df):
        """Formatlanmış Excel dosyası oluşturur"""
        try:
            output = BytesIO()
            
            # Workbook oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "CashBack Analizi"
            
            # Başlık satırı
            headers = ['Müşteri Kimliği', 'Müşteri Adı', 'İşlem Adedi', 'Toplam Miktar (₺)']
            ws.append(headers)
            
            # Veri satırları
            for _, row in df.iterrows():
                ws.append([
                    int(row['Müşteri_Kimliği']),
                    row['Müşteri_Adı'],
                    int(row['Adet']),
                    float(row['Toplam_Miktar'])
                ])
            
            # Stil tanımlamaları
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            data_font = Font(color="000000")
            data_fill_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
            data_fill_2 = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            alignment = Alignment(horizontal='center', vertical='center')
            
            # Başlık formatı
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = border
            
            # Veri formatı
            for row in range(2, len(df) + 2):
                fill = data_fill_1 if row % 2 == 0 else data_fill_2
                
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row, column=col)
                    cell.font = data_font
                    cell.fill = fill
                    cell.alignment = alignment
                    cell.border = border
                    
                    # Sayı formatı
                    if col == 4:  # Toplam Miktar sütunu
                        cell.number_format = '₺#,##0.00'
            
            # Sütun genişliklerini ayarla
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 20
            
            # AutoFilter ekle
            ws.auto_filter.ref = f"A1:D{len(df) + 1}"
            
            # Dondurulan pencere
            ws.freeze_panes = "A2"
            
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            st.error(f"❌ Excel dosyası oluşturma hatası: {str(e)}")
            return None
    
    def create_historical_analysis_excel(self, data, date_range):
        """Tarihsel analiz için renkli ve çarpıcı Excel dosyası oluşturur"""
        try:
            output = BytesIO()
            
            # DataFrame oluştur ve sırala
            df = pd.DataFrame(data)
            if not df.empty:
                df = df.sort_values('Toplam_Miktar', ascending=False).reset_index(drop=True)
            
            # Workbook oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Tarihsel CashBack Analizi"
            
            # Başlık ve tarih aralığı
            start_date = date_range[0].strftime('%d.%m.%Y')
            end_date = date_range[1].strftime('%d.%m.%Y')
            title = f"CashBack Analizi ({start_date} - {end_date})"
            
            # Ana başlık (A1-E2 birleştir)
            ws.merge_cells('A1:E2')
            title_cell = ws['A1']
            title_cell.value = title
            title_cell.font = Font(bold=True, size=16, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="1f4e79", end_color="1f4e79", fill_type="solid")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Özet istatistikler (4. satır)
            ws['A4'] = "📊 ÖZET İSTATİSTİKLER"
            ws.merge_cells('A4:E4')
            summary_cell = ws['A4']
            summary_cell.font = Font(bold=True, size=12, color="FFFFFF")
            summary_cell.fill = PatternFill(start_color="2e75b6", end_color="2e75b6", fill_type="solid")
            summary_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # İstatistik değerleri
            total_customers = len(df)
            total_amount = df['Toplam_Miktar'].sum() if not df.empty else 0
            total_transactions = df['Adet'].sum() if not df.empty else 0
            avg_per_customer = total_amount / total_customers if total_customers > 0 else 0
            
            stats = [
                ['Toplam Müşteri:', f'{total_customers:,}'],
                ['Toplam Miktar:', f'₺{total_amount:,.2f}'],
                ['Toplam İşlem:', f'{total_transactions:,}'],
                ['Müşteri Başına Ort.:', f'₺{avg_per_customer:,.2f}']
            ]
            
            for i, (label, value) in enumerate(stats):
                row = 5 + i
                ws[f'B{row}'] = label
                ws[f'C{row}'] = value
                
                # Stil
                ws[f'B{row}'].font = Font(bold=True, color="2e75b6")
                ws[f'C{row}'].font = Font(bold=True, color="c55a11")
                ws[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
                ws[f'C{row}'].alignment = Alignment(horizontal='left', vertical='center')
            
            # Veri tablosu başlığı
            data_start_row = 10
            ws[f'A{data_start_row}'] = "📋 DETAYLI MÜŞTERI ANALİZİ"
            ws.merge_cells(f'A{data_start_row}:E{data_start_row}')
            data_title_cell = ws[f'A{data_start_row}']
            data_title_cell.font = Font(bold=True, size=12, color="FFFFFF")
            data_title_cell.fill = PatternFill(start_color="70ad47", end_color="70ad47", fill_type="solid")
            data_title_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Sütun başlıkları
            headers = ['Sıra', 'Müşteri Kimliği', 'Müşteri Adı', 'İşlem Adedi', 'Toplam Miktar (₺)']
            header_row = data_start_row + 1
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF", size=11)
                cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(
                    left=Side(style='medium'),
                    right=Side(style='medium'),
                    top=Side(style='medium'),
                    bottom=Side(style='medium')
                )
            
            # Veri satırları
            if not df.empty:
                for idx, (_, row) in enumerate(df.iterrows()):
                    excel_row = header_row + 1 + idx
                    
                    # Renk gradasyonu (en yüksek yeşil, en düşük sarı)
                    if idx < len(df) * 0.2:  # En yüksek %20
                        fill_color = "d5e8d4"  # Açık yeşil
                        font_color = "2d5016"
                    elif idx < len(df) * 0.5:  # Orta %30
                        fill_color = "fff2cc"  # Açık sarı
                        font_color = "7f6000"
                    else:  # Alt %50
                        fill_color = "f8cecc"  # Açık kırmızı
                        font_color = "b85450"
                    
                    values = [
                        idx + 1,
                        int(row['Müşteri_Kimliği']),
                        row['Müşteri_Adı'],
                        int(row['Adet']),
                        float(row['Toplam_Miktar'])
                    ]
                    
                    for col, value in enumerate(values, 1):
                        cell = ws.cell(row=excel_row, column=col)
                        cell.value = value
                        
                        # Sayı formatı
                        if col == 5:  # Toplam Miktar sütunu
                            cell.number_format = '₺#,##0.00'
                        elif col in [1, 2, 4]:  # Sayı sütunları
                            cell.number_format = '#,##0'
                        
                        # Stil
                        cell.font = Font(color=font_color, size=10)
                        cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        cell.border = Border(
                            left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin')
                        )
            
            # Sütun genişliklerini ayarla
            ws.column_dimensions['A'].width = 8   # Sıra
            ws.column_dimensions['B'].width = 18  # Müşteri Kimliği
            ws.column_dimensions['C'].width = 35  # Müşteri Adı
            ws.column_dimensions['D'].width = 15  # İşlem Adedi
            ws.column_dimensions['E'].width = 20  # Toplam Miktar
            
            # AutoFilter ekle (sadece veri tablosuna)
            if not df.empty:
                ws.auto_filter.ref = f"A{header_row}:E{header_row + len(df)}"
            
            # Dondurulan pencere
            ws.freeze_panes = f"A{header_row + 1}"
            
            # Sayfa düzeni
            ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 0
            
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            st.error(f"❌ Tarihsel Excel dosyası oluşturma hatası: {str(e)}")
            return None

class DataManager:
    def __init__(self):
        self.json_file = "CashBack.json"
    
    def save_to_json(self, df, selected_date=None):
        """DataFrame'i JSON dosyasına belirli tarihe kaydeder"""
        try:
            if selected_date is None:
                selected_date = date.today()
            
            # Tarihi string formatına çevir
            date_str = selected_date.strftime("%Y-%m-%d")
            
            # Mevcut verileri yükle
            existing_data = self.load_all_data()
            
            # Aynı tarihli kayıt varsa güncelle, yoksa yeni ekle
            updated = False
            for entry in existing_data:
                if entry.get("date", "").split("_")[0] == date_str:
                    entry["data"] = df.to_dict('records')
                    entry["timestamp"] = datetime.now().isoformat()
                    updated = True
                    break
            
            if not updated:
                # Yeni veriyi ekle
                new_data = {
                    "date": f"{date_str}_{datetime.now().strftime('%H:%M:%S')}",
                    "timestamp": datetime.now().isoformat(),
                    "data": df.to_dict('records')
                }
                existing_data.append(new_data)
            
            # Tarihe göre sırala (en yeni en üstte)
            existing_data.sort(key=lambda x: x.get("timestamp", ""), reverse=True)
            
            # Dosyaya kaydet
            with open(self.json_file, "w", encoding="utf-8") as f:
                json.dump(existing_data, f, indent=4, ensure_ascii=False)
            
            return True
            
        except Exception as e:
            st.error(f"❌ JSON kaydetme hatası: {str(e)}")
            return False
    
    def load_all_data(self):
        """JSON dosyasından tüm verileri yükler"""
        try:
            if os.path.exists(self.json_file):
                with open(self.json_file, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    
                    if isinstance(data, list) and data and isinstance(data[0], dict):
                        # Yeni format kontrolü
                        if "date" in data[0]:
                            return data
                        else:
                            # Eski format - yeni formata çevir
                            return [{
                                "date": datetime.now().strftime("%Y-%m-%d_%H:%M:%S"),
                                "timestamp": datetime.now().isoformat(),
                                "data": data
                            }]
                    
                    return data if isinstance(data, list) else []
            
            return []
            
        except Exception as e:
            st.error(f"❌ JSON dosyası okuma hatası: {str(e)}")
            return []
    
    def get_data_by_date_range(self, start_date, end_date):
        """Tarih aralığına göre verileri filtreler"""
        try:
            all_data = self.load_all_data()
            
            if not all_data:
                return []
            
            filtered_data = []
            
            for entry in all_data:
                # Tarih stringinden tarihi çıkar
                date_str = entry.get("date", "").split("_")[0]
                if date_str:
                    try:
                        entry_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        if start_date <= entry_date <= end_date:
                            filtered_data.extend(entry["data"])
                    except ValueError:
                        continue
            
            return filtered_data
            
        except Exception as e:
            st.error(f"❌ Tarihsel veri yükleme hatası: {str(e)}")
            return []
    
    def get_last_7_days_data(self):
        """Son 7 günün verilerini getirir"""
        end_date = date.today()
        start_date = end_date - timedelta(days=7)
        return self.get_data_by_date_range(start_date, end_date)
    
    def get_monthly_data(self, year=None, month=None):
        """Belirli bir ayın verilerini getirir"""
        if year is None:
            year = date.today().year
        if month is None:
            month = date.today().month
        
        # Ayın ilk ve son günü
        start_date = date(year, month, 1)
        last_day = calendar.monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        
        return self.get_data_by_date_range(start_date, end_date)
    
    def get_daily_totals(self, start_date, end_date):
        """Günlük toplam miktarları getirir"""
        try:
            all_data = self.load_all_data()
            daily_totals = {}
            
            for entry in all_data:
                date_str = entry.get("date", "").split("_")[0]
                if date_str:
                    try:
                        entry_date = datetime.strptime(date_str, "%Y-%m-%d").date()
                        if start_date <= entry_date <= end_date:
                            if entry_date not in daily_totals:
                                daily_totals[entry_date] = 0
                            
                            for record in entry["data"]:
                                daily_totals[entry_date] += record.get("Toplam_Miktar", 0)
                    except ValueError:
                        continue
            
            return daily_totals
            
        except Exception as e:
            st.error(f"❌ Günlük toplam hesaplama hatası: {str(e)}")
            return {}

class Visualizer:
    def __init__(self):
        pass
    
    def create_top_customers_chart(self, data, title="En Fazla CashBack Alan Müşteriler", top_n=10):
        """En fazla cashback alan müşteriler için pasta grafik oluşturur"""
        try:
            if not data:
                return None
            
            df = pd.DataFrame(data)
            
            # En yüksek miktarlı müşterileri seç
            top_customers = df.nlargest(top_n, 'Toplam_Miktar')
            
            # Pasta grafik oluştur
            fig = px.pie(
                top_customers, 
                values='Toplam_Miktar', 
                names='Müşteri_Adı',
                title=title,
                hole=0.4,  # Donut chart yapar
                color_discrete_sequence=px.colors.qualitative.Set3
            )
            
            fig.update_traces(
                textposition='inside', 
                textinfo='percent+label',
                hovertemplate='<b>%{label}</b><br>Miktar: ₺%{value:,.0f}<br>Oran: %{percent}<extra></extra>'
            )
            
            fig.update_layout(
                showlegend=True,
                height=500,
                font=dict(size=12)
            )
            
            return fig
            
        except Exception as e:
            st.error(f"❌ Grafik oluşturma hatası: {str(e)}")
            return None
    
    def create_daily_trend_chart(self, daily_data, title="Günlük CashBack Trendi"):
        """Günlük CashBack trendini gösteren çizgi grafik"""
        try:
            if not daily_data:
                return None
            
            dates = list(daily_data.keys())
            amounts = list(daily_data.values())
            
            fig = go.Figure()
            
            fig.add_trace(go.Scatter(
                x=dates,
                y=amounts,
                mode='lines+markers',
                name='Günlük CashBack',
                line=dict(color='#1f77b4', width=3),
                marker=dict(size=8, color='#ff7f0e'),
                hovertemplate='<b>Tarih:</b> %{x}<br><b>Miktar:</b> ₺%{y:,.0f}<extra></extra>'
            ))
            
            fig.update_layout(
                title=title,
                xaxis_title="Tarih",
                yaxis_title="CashBack Miktarı (₺)",
                height=400,
                showlegend=False,
                hovermode='x unified'
            )
            
            return fig
            
        except Exception as e:
            st.error(f"❌ Trend grafik oluşturma hatası: {str(e)}")
            return None
    
    def create_top_customers_bar_chart(self, data, title="En Aktif Müşteriler", top_n=15):
        """En aktif müşteriler için bar chart"""
        try:
            if not data:
                return None
            
            df = pd.DataFrame(data)
            top_customers = df.nlargest(top_n, 'Toplam_Miktar')
            
            fig = px.bar(
                top_customers,
                x='Toplam_Miktar',
                y='Müşteri_Adı',
                orientation='h',
                title=title,
                color='Toplam_Miktar',
                color_continuous_scale='viridis',
                text='Toplam_Miktar'
            )
            
            fig.update_traces(
                texttemplate='₺%{text:,.0f}',
                textposition='inside',
                hovertemplate='<b>%{y}</b><br>Miktar: ₺%{x:,.0f}<extra></extra>'
            )
            
            fig.update_layout(
                height=600,
                yaxis={'categoryorder': 'total ascending'},
                showlegend=False
            )
            
            return fig
            
        except Exception as e:
            st.error(f"❌ Bar grafik oluşturma hatası: {str(e)}")
            return None

def main():
    # Ana başlık ve tarih gösterimi
    col1, col2, col3 = st.columns([2, 1, 1])
    
    with col1:
        st.title("💰 CashBack Düzeltmesi Analizi")
    
    with col2:
        st.metric("📅 Bugünün Tarihi", date.today().strftime("%d.%m.%Y"))
    
    with col3:
        current_time = datetime.now().strftime("%H:%M")
        st.metric("🕐 Şu Anki Saat", current_time)
    
    # Veri yönetimi ve görselleştirme sınıfları
    data_manager = DataManager()
    visualizer = Visualizer()
    
    # Dashboard - Üst kısım istatistikleri
    st.markdown("---")
    st.subheader("📊 Dashboard")
    
    # Son 7 günlük veriler
    last_7_days_data = data_manager.get_last_7_days_data()
    last_7_days_total = sum(record.get('Toplam_Miktar', 0) for record in last_7_days_data)
    
    # Aylık veriler
    monthly_data = data_manager.get_monthly_data()
    monthly_total = sum(record.get('Toplam_Miktar', 0) for record in monthly_data)
    
    # Son 30 günlük trendler (daha fazla gün göstermek için)
    end_date = date.today()
    start_date = end_date - timedelta(days=30)
    daily_totals = data_manager.get_daily_totals(start_date, end_date)
    
    # Günlük CashBack detayları
    st.subheader("📅 Günlük CashBack Analizi")
    
    if daily_totals:
        # Günlük verileri tarihe göre sırala (en yeni en üstte)
        sorted_daily = dict(sorted(daily_totals.items(), key=lambda x: x[0], reverse=True))
        
        # Her gün için metrik göster
        days_list = list(sorted_daily.items())
        
        # İlk 7 günü göster
        for i in range(0, min(len(days_list), 7), 4):
            cols = st.columns(4)
            for j, col in enumerate(cols):
                if i + j < len(days_list):
                    date_obj, amount = days_list[i + j]
                    # Türkçe ay isimleri
                    turkish_months = {
                        1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
                        5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
                        9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
                    }
                    
                    day = date_obj.day
                    month = turkish_months[date_obj.month]
                    
                    with col:
                        st.metric(
                            f"📅 {day} {month}",
                            f"₺{amount:,.0f}",
                            delta="CashBack"
                        )
    else:
        st.info("📊 Henüz günlük veri bulunmuyor")
    
    # Aylık toplam metrikleri
    col1, col2, col3 = st.columns(3)
    
    with col1:
        current_month = date.today().strftime("%B %Y")
        st.metric(
            f"📅 Bu Ay Toplam",
            f"₺{monthly_total:,.0f}",
            delta=f"{len(monthly_data)} işlem"
        )
    
    with col2:
        avg_daily = last_7_days_total / 7 if last_7_days_total > 0 else 0
        st.metric(
            "💡 Günlük Ortalama",
            f"₺{avg_daily:,.0f}",
            delta="Son 7 gün"
        )
    
    with col3:
        if last_7_days_data:
            unique_customers = len(set(record.get('Müşteri_Kimliği') for record in last_7_days_data))
            st.metric(
                "👥 Aktif Müşteri",
                f"{unique_customers}",
                delta="Son 7 gün"
            )
    
    # Görselleştirmeler
    if last_7_days_data or monthly_data:
        col1, col2 = st.columns(2)
        
        with col1:
            # Günlük trend grafiği
            if daily_totals:
                trend_chart = visualizer.create_daily_trend_chart(daily_totals)
                if trend_chart:
                    st.plotly_chart(trend_chart, use_container_width=True)
        
        with col2:
            # En fazla cashback alan müşteriler pasta grafik
            display_data = monthly_data if monthly_data else last_7_days_data
            if display_data:
                pie_chart = visualizer.create_top_customers_chart(display_data)
                if pie_chart:
                    st.plotly_chart(pie_chart, use_container_width=True)
        
        # Bar chart için tam genişlik
        if monthly_data:
            bar_chart = visualizer.create_top_customers_bar_chart(monthly_data)
            if bar_chart:
                st.plotly_chart(bar_chart, use_container_width=True)
    
    st.markdown("---")
    
    # Sidebar
    with st.sidebar:
        st.header("⚙️ İşlemler")
        
        # Excel dosyası yükleme
        st.subheader("📁 Dosya Yükleme")
        uploaded_file = st.file_uploader(
            "Excel dosyasını seçin",
            type=['xlsx', 'xls'],
            help="CashBack Düzeltmesi verilerini içeren Excel dosyasını yükleyin"
        )
        
        # Tarih seçici
        st.subheader("📅 Kayıt Tarihi")
        st.info("⚠️ Önce tarihi seçin, sonra verileri yükleyin!")
        selected_date = st.date_input(
            "📅 Hangi tarihe kaydetmek istiyorsunuz?",
            value=date.today(),
            help="UYARI: Seçilen tarihe veriler kaydedilecek. Önce tarihi değiştirin!"
        )
        st.warning(f"🎯 Seçilen kayıt tarihi: **{selected_date.strftime('%d.%m.%Y')}**")
        
        # Tarihsel veri görüntüleme
        st.subheader("📈 Tarihsel Analiz")
        
        # Tarih aralığı seçici
        date_range = st.date_input(
            "Analiz için tarih aralığı seçin",
            value=[date.today() - timedelta(days=7), date.today()],
            help="Belirtilen tarih aralığındaki veriler analiz edilecek"
        )
        
        if len(date_range) == 2:
            if st.button("📊 Tarihsel Analiz Yap"):
                historical_data = data_manager.get_data_by_date_range(date_range[0], date_range[1])
                if historical_data:
                    st.success(f"✅ {len(historical_data)} kayıt bulundu")
                    
                    # Tarihsel verileri ana sayfada göster
                    st.session_state['historical_data'] = historical_data
                    st.session_state['historical_range'] = date_range
                    
                    # Excel indirme butonu
                    processor = ExcelProcessor()
                    excel_data = processor.create_historical_analysis_excel(historical_data, date_range)
                    if excel_data:
                        file_name = f"Tarihsel_Analiz_{date_range[0].strftime('%Y%m%d')}_{date_range[1].strftime('%Y%m%d')}.xlsx"
                        st.download_button(
                            label="📥 Tarihsel Analizi Excel İndir",
                            data=excel_data,
                            file_name=file_name,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            type="primary"
                        )
                else:
                    st.warning("⚠️ Belirtilen tarih aralığında veri bulunamadı")
    
    # Ana içerik alanı
    if uploaded_file is not None:
        try:
            # Excel işleme
            processor = ExcelProcessor()
            
            # Dosyayı oku
            if uploaded_file.name.endswith('.xlsx'):
                df = pd.read_excel(uploaded_file, engine='openpyxl')
            else:
                df = pd.read_excel(uploaded_file)
            
            st.info(f"📋 Dosya yüklendi: {uploaded_file.name}")
            st.info(f"📊 Toplam satır sayısı: {len(df)}")
            
            # Verileri işle
            processed_df = processor.process_cashback_data(df)
            
            if not processed_df.empty:
                # İşlenmiş verileri göster
                st.subheader("📋 İşlenmiş Veriler")
                st.dataframe(processed_df, use_container_width=True)
                
                # Özet istatistikler
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    total_customers = len(processed_df)
                    st.metric("👥 Toplam Müşteri", total_customers)
                
                with col2:
                    total_transactions = processed_df['Adet'].sum()
                    st.metric("📊 Toplam İşlem", f"{total_transactions:,}")
                
                with col3:
                    total_amount = processed_df['Toplam_Miktar'].sum()
                    st.metric("💰 Toplam Miktar", f"₺{total_amount:,.2f}")
                
                # Kaydetme ve indirme işlemleri
                col1, col2 = st.columns(2)
                
                with col1:
                    if st.button("💾 JSON'a Kaydet", type="primary"):
                        if data_manager.save_to_json(processed_df, selected_date):
                            st.success(f"✅ Veriler {selected_date.strftime('%d.%m.%Y')} tarihine kaydedildi!")
                            st.rerun()
                        else:
                            st.error("❌ Kaydetme işlemi başarısız!")
                
                with col2:
                    # Excel indirme
                    excel_data = processor.create_formatted_excel(processed_df)
                    if excel_data:
                        st.download_button(
                            label="📥 Excel İndir",
                            data=excel_data,
                            file_name=f"CashBack_Analizi_{selected_date.strftime('%Y%m%d')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                # Görselleştirmeler - yeni yüklenen veriler için
                st.subheader("📊 Analiz Grafikleri")
                
                col1, col2 = st.columns(2)
                
                with col1:
                    # Pasta grafik
                    pie_chart = visualizer.create_top_customers_chart(
                        processed_df.to_dict('records'),
                        "Yeni Yüklenen Veriler - En Fazla CashBack"
                    )
                    if pie_chart:
                        st.plotly_chart(pie_chart, use_container_width=True)
                
                with col2:
                    # Bar chart
                    bar_chart = visualizer.create_top_customers_bar_chart(
                        processed_df.to_dict('records'),
                        "Yeni Yüklenen Veriler - Müşteri Sıralaması"
                    )
                    if bar_chart:
                        st.plotly_chart(bar_chart, use_container_width=True)
        
        except Exception as e:
            st.error(f"❌ Dosya işleme hatası: {str(e)}")
            st.error(f"📋 Hata detayı: {traceback.format_exc()}")
    
    # Tarihsel veri analizi sonuçları
    if 'historical_data' in st.session_state:
        st.markdown("---")
        st.subheader(f"📈 Tarihsel Analiz Sonuçları ({st.session_state['historical_range'][0]} - {st.session_state['historical_range'][1]})")
        
        historical_data = st.session_state['historical_data']
        historical_df = pd.DataFrame(historical_data)
        
        if not historical_df.empty:
            # Müşteri bazında gruplama
            grouped_historical = historical_df.groupby(['Müşteri_Kimliği', 'Müşteri_Adı']).agg({
                'Adet': 'sum',
                'Toplam_Miktar': 'sum'
            }).reset_index()
            
            grouped_historical = grouped_historical.sort_values('Toplam_Miktar', ascending=False).reset_index(drop=True)
            
            # Tarihsel verileri göster
            st.dataframe(grouped_historical, use_container_width=True)
            
            # Tarihsel özet ve indirme
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                st.metric("👥 Müşteri Sayısı", len(grouped_historical))
            
            with col2:
                st.metric("📊 Toplam İşlem", f"{grouped_historical['Adet'].sum():,}")
            
            with col3:
                st.metric("💰 Toplam Miktar", f"₺{grouped_historical['Toplam_Miktar'].sum():,.2f}")
            
            with col4:
                # Excel indirme butonu
                processor = ExcelProcessor()
                excel_data = processor.create_historical_analysis_excel(
                    grouped_historical.to_dict('records'), 
                    st.session_state['historical_range']
                )
                if excel_data:
                    file_name = f"Tarihsel_Analiz_{st.session_state['historical_range'][0].strftime('%Y%m%d')}_{st.session_state['historical_range'][1].strftime('%Y%m%d')}.xlsx"
                    st.download_button(
                        label="📥 Excel İndir",
                        data=excel_data,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
            
            # Tarihsel grafikler
            col1, col2 = st.columns(2)
            
            with col1:
                pie_chart = visualizer.create_top_customers_chart(
                    grouped_historical.to_dict('records'),
                    "Tarihsel Veriler - En Fazla CashBack"
                )
                if pie_chart:
                    st.plotly_chart(pie_chart, use_container_width=True)
            
            with col2:
                bar_chart = visualizer.create_top_customers_bar_chart(
                    grouped_historical.to_dict('records'),
                    "Tarihsel Veriler - Müşteri Sıralaması"
                )
                if bar_chart:
                    st.plotly_chart(bar_chart, use_container_width=True)
    
    # Alt bilgi
    st.markdown("---")
    st.markdown(
        """
        <div style='text-align: center; color: gray;'>
        💰 CashBack Düzeltmesi Analizi | Gelişmiş Veri Analizi ve Görselleştirme
        </div>
        """,
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
